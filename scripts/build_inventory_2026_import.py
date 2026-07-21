from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps


DEFAULT_SOURCE_NAME = "BASE DE DATOS 2026.xlsx"
IMPORT_KEY = "base_datos_2026_v1"
EMU_PER_POINT = 12700


@dataclass
class ImageCandidate:
    data: bytes
    content_hash: str
    display_area: float
    mapped_row: int


@dataclass
class SourceRow:
    sheet: str
    row: int
    warehouse: str
    snapshot_date: dt.date | None
    source_priority: int
    numero_parte: str
    descripcion: str
    marca: str
    categoria: str
    stock: int | None
    stock_raw: str
    image: ImageCandidate | None = None

    @property
    def source_ref(self) -> str:
        return f"{self.sheet}!{self.row}"

    @property
    def completeness(self) -> int:
        return sum(bool(value) for value in (self.numero_parte, self.marca, self.categoria, self.image))


@dataclass
class Product:
    source_key: str
    numero_parte: str
    descripcion: str
    marca: str
    categoria: str
    almacen: str
    stock: int
    stock_known: bool
    fotografia: str | None
    sources: list[SourceRow] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


class UnionFind:
    def __init__(self, size: int) -> None:
        self.parent = list(range(size))
        self.rank = [0] * size

    def find(self, value: int) -> int:
        while self.parent[value] != value:
            self.parent[value] = self.parent[self.parent[value]]
            value = self.parent[value]
        return value

    def union(self, left: int, right: int) -> None:
        left_root = self.find(left)
        right_root = self.find(right)
        if left_root == right_root:
            return
        if self.rank[left_root] < self.rank[right_root]:
            left_root, right_root = right_root, left_root
        self.parent[right_root] = left_root
        if self.rank[left_root] == self.rank[right_root]:
            self.rank[left_root] += 1


def normalize(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).replace("\r", " ").replace("\n", " ").replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def ascii_fold(value: Any) -> str:
    text = unicodedata.normalize("NFKD", normalize(value))
    return text.encode("ascii", "ignore").decode("ascii")


def canonical(value: Any) -> str:
    text = ascii_fold(value).upper()
    text = text.replace("\\", "/").replace("×", "X").replace("*", "X")
    text = text.replace("”", '"').replace("“", '"').replace("''", '"')
    text = text.replace('"', "")
    text = re.sub(r"\s*([/X#\-])\s*", r"\1", text)
    text = re.sub(r"[^A-Z0-9/#.\-\"]+", " ", text)
    return re.sub(r"\s+", " ", text).strip(" .-")


def header_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", ascii_fold(value).upper()).strip()


def measurement_signature(value: str) -> tuple[str, ...]:
    text = canonical(value)
    tokens = re.findall(r"[A-Z]*\d+(?:[./-]\d+)*(?:[A-Z]+)?|\d+/\d+", text)
    return tuple(tokens)


def compatible_brand(left: str, right: str) -> bool:
    left_key = canonical(left)
    right_key = canonical(right)
    if not left_key or not right_key:
        return True
    if left_key == right_key:
        return True
    return SequenceMatcher(None, left_key, right_key).ratio() >= 0.97


def normalize_warehouse(sheet_name: str) -> str:
    key = canonical(sheet_name)
    if key == "INV.PLAFONES":
        return "Almacén General - Plafones"
    if key.startswith("INV.DISPONIBLE"):
        return "Almacén General - Disponible"
    if key == "INV.SUSPENSION MECANICA":
        return "Suspensión mecánica"
    if key == "INV.LAMINAS":
        return "Almacén General - Láminas"
    if key == "INV.TUBOS MONTERREY":
        return "Almacén General - Tubería"
    if key == "INV.CILINDROS LINDE":
        return "Cilindros Linde"
    if key in {"INV. ALMACEN PINTURA RACSA Y NE", "INVENTARIO PINTURA"}:
        return "Almacén Pintura"
    if key == "INV.ALMACEN GENERAL":
        return "Almacén General"
    if key == "INV. ALMACEN PINTURA":
        return "Almacén Pintura"
    if key == "INV.GENERAL CONEXION":
        return "Almacén General - Conexiones"
    if key == "INV. ALMACEN SUSPENSIONES":
        return "Almacén Suspensiones"
    if key.startswith("INV.ALMACEN 2") or key == "INV.HORNO PINTURA":
        return "Almacén 2"
    return re.sub(r"^INV(?:ENTARIO)?[. ]*", "", normalize(sheet_name), flags=re.IGNORECASE).strip() or "Almacén principal"


def sheet_priority(sheet_name: str) -> int:
    key = canonical(sheet_name)
    if key == "INV.ALMACEN 2":
        return 40
    if key == "INV.HORNO PINTURA":
        return 30
    if key == "INV.ALMACEN 2 (2)":
        return 20
    if key == "INV.ALMACEN 2 (3)":
        return 10
    if key == "INVENTARIO PINTURA":
        return 35
    return 25


def clean_category(value: str, warehouse: str) -> str:
    category = normalize(value).upper()
    category_key = canonical(category)
    replacements = {
        "LLENADO POR LE FONDO": "LLENADO POR EL FONDO",
        "LLENADO POR EL FONDO": "LLENADO POR EL FONDO",
    }
    if category_key in replacements:
        return replacements[category_key]
    if not re.search(r"[A-Z]", ascii_fold(category).upper()):
        if "Tubería" in warehouse:
            return "TUBERIA"
        return "SIN CATEGORÍA"
    return category or "SIN CATEGORÍA"


def parse_stock(value: Any) -> tuple[int | None, str]:
    raw = normalize(value)
    if value is None or not raw or raw.startswith("#"):
        return None, raw
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if not math.isfinite(float(value)):
            return None, raw
        return max(0, int(round(float(value)))), raw
    if canonical(raw) in {"N/A", "NA", "SIN STOCK", "NO DISPONIBLE", "PENDIENTE"}:
        return None, raw
    match = re.search(r"-?\d+(?:[.,]\d+)?", raw.replace(",", ""))
    if not match:
        return None, raw
    return max(0, int(round(float(match.group(0))))), raw


def find_header(ws: Any) -> tuple[int, dict[str, int]] | None:
    for row_number in range(1, min(ws.max_row, 30) + 1):
        keys = [header_key(ws.cell(row_number, column).value) for column in range(1, ws.max_column + 1)]
        if "DESCRIPCION" not in keys or "STOCK" not in keys:
            continue
        mapping: dict[str, int] = {}
        for column, key in enumerate(keys, start=1):
            if key in {"NO PARTE", "NO DE PARTE", "CLAVE INTERNA", "NO PARTE CLAVE INTERNA"}:
                mapping["numero_parte"] = column
            elif key == "DESCRIPCION":
                mapping["descripcion"] = column
            elif key == "MARCA":
                mapping["marca"] = column
            elif key == "CATEGORIA":
                mapping["categoria"] = column
            elif key == "STOCK":
                mapping["stock"] = column
            elif key in {"APOYO VISUAL", "FOTOGRAFIA", "IMAGEN"}:
                mapping["fotografia"] = column
        return row_number, mapping
    return None


def extract_snapshot_date(ws: Any) -> dt.date | None:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
        for value in row:
            if isinstance(value, dt.datetime):
                return value.date()
            if isinstance(value, dt.date):
                return value
            text = normalize(value)
            match = re.search(r"\b(\d{1,2})/(\d{1,2})/(20\d{2})\b", text)
            if match:
                try:
                    return dt.date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
                except ValueError:
                    pass
    return None


def row_height_points(ws: Any, row_number: int) -> float:
    return float(ws.row_dimensions[row_number].height or ws.sheet_format.defaultRowHeight or 15.0)


def marker_y_points(ws: Any, marker: Any) -> float:
    before = sum(row_height_points(ws, row_number) for row_number in range(1, marker.row + 1))
    return before + float(marker.rowOff or 0) / EMU_PER_POINT


def displayed_vertical_bounds(ws: Any, image: Any) -> tuple[float, float, float]:
    anchor = image.anchor
    start = marker_y_points(ws, anchor._from)
    if hasattr(anchor, "ext") and anchor.ext is not None:
        height = float(anchor.ext.cy) / EMU_PER_POINT
        width = float(anchor.ext.cx) / EMU_PER_POINT
        return start, start + height, max(1.0, height * width)
    if hasattr(anchor, "to") and anchor.to is not None:
        end = marker_y_points(ws, anchor.to)
        low, high = sorted((start, end))
        height = max(1.0, high - low)
        native_ratio = max(0.2, min(5.0, float(image.width or 1) / float(image.height or 1)))
        return low, high, height * height * native_ratio
    height = float(image.height or 1) * 0.75
    width = float(image.width or 1) * 0.75
    return start, start + height, max(1.0, height * width)


def row_center_points(ws: Any, row_number: int) -> float:
    before = sum(row_height_points(ws, number) for number in range(1, row_number))
    return before + row_height_points(ws, row_number) / 2


def image_column(image: Any) -> int:
    return int(image.anchor._from.col) + 1


def map_images(ws: Any, source_rows: dict[int, SourceRow], photo_column: int | None) -> dict[int, ImageCandidate]:
    if not source_rows:
        return {}
    candidates: dict[int, list[ImageCandidate]] = defaultdict(list)
    row_centers = {row: row_center_points(ws, row) for row in source_rows}

    for image in getattr(ws, "_images", []):
        if photo_column is not None and abs(image_column(image) - photo_column) > 1:
            continue
        try:
            low, high, area = displayed_vertical_bounds(ws, image)
            center = (low + high) / 2
            nearest_row = min(row_centers, key=lambda row: abs(row_centers[row] - center))
            tolerance = max(45.0, row_height_points(ws, nearest_row) * 0.8)
            if abs(row_centers[nearest_row] - center) > tolerance:
                continue
            data = image._data()
            with Image.open(BytesIO(data)) as probe:
                probe.verify()
            candidate = ImageCandidate(
                data=data,
                content_hash=hashlib.sha1(data).hexdigest(),
                display_area=area,
                mapped_row=nearest_row,
            )
            candidates[nearest_row].append(candidate)
        except Exception:
            continue

    selected: dict[int, ImageCandidate] = {}
    for row, options in candidates.items():
        selected[row] = max(options, key=lambda option: option.display_area)
    return selected


def read_source_rows(workbook: Any) -> tuple[list[SourceRow], list[dict[str, Any]]]:
    source_rows: list[SourceRow] = []
    skipped: list[dict[str, Any]] = []

    for ws in workbook.worksheets:
        special_paint = canonical(ws.title) == "INVENTARIO PINTURA"
        header = find_header(ws)
        if header is None and not special_paint:
            skipped.append({"sheet": ws.title, "row": None, "reason": "No se encontró encabezado de descripción y stock"})
            continue

        if special_paint:
            header_row = 0
            mapping = {"descripcion": 2, "marca": 3, "stock": 4, "fotografia": 5}
        else:
            assert header is not None
            header_row, mapping = header

        rows_for_sheet: dict[int, SourceRow] = {}
        snapshot_date = extract_snapshot_date(ws)
        for row_number in range(header_row + 1, ws.max_row + 1):
            descripcion = normalize(ws.cell(row_number, mapping["descripcion"]).value)
            if not descripcion or not canonical(descripcion):
                continue
            if canonical(descripcion) in {"DESCRIPCION", "TOTAL", "TOTALES", "APOYO VISUAL"}:
                continue
            numero_parte = normalize(ws.cell(row_number, mapping.get("numero_parte", 0)).value) if mapping.get("numero_parte") else ""
            marca = normalize(ws.cell(row_number, mapping.get("marca", 0)).value) if mapping.get("marca") else ""
            categoria = normalize(ws.cell(row_number, mapping.get("categoria", 0)).value) if mapping.get("categoria") else ""
            if special_paint and not categoria:
                categoria = "PINTURA"
            stock_value = ws.cell(row_number, mapping["stock"]).value
            stock, stock_raw = parse_stock(stock_value)
            rows_for_sheet[row_number] = SourceRow(
                sheet=ws.title.strip(),
                row=row_number,
                warehouse=normalize_warehouse(ws.title),
                snapshot_date=snapshot_date,
                source_priority=sheet_priority(ws.title),
                numero_parte=numero_parte,
                descripcion=descripcion,
                marca=marca,
                categoria=categoria,
                stock=stock,
                stock_raw=stock_raw,
            )

        images = map_images(ws, rows_for_sheet, mapping.get("fotografia"))
        for row_number, record in rows_for_sheet.items():
            record.image = images.get(row_number)
            source_rows.append(record)

    return source_rows, skipped


def should_merge(left: SourceRow, right: SourceRow) -> bool:
    if not compatible_brand(left.marca, right.marca):
        return False
    left_desc = canonical(left.descripcion)
    right_desc = canonical(right.descripcion)
    if left_desc == right_desc:
        return True
    left_signature = measurement_signature(left.descripcion)
    right_signature = measurement_signature(right.descripcion)
    if left_signature != right_signature:
        return False
    ratio = SequenceMatcher(None, left_desc, right_desc).ratio()
    if ratio >= 0.985:
        return True
    left_part = canonical(left.numero_parte)
    right_part = canonical(right.numero_parte)
    return bool(left_part and left_part == right_part and ratio >= 0.90)


def cluster_rows(rows: list[SourceRow]) -> list[list[SourceRow]]:
    union_find = UnionFind(len(rows))
    exact_description: dict[str, list[int]] = defaultdict(list)
    for index, row in enumerate(rows):
        exact_description[canonical(row.descripcion)].append(index)

    for indices in exact_description.values():
        for position, left_index in enumerate(indices):
            for right_index in indices[position + 1 :]:
                if should_merge(rows[left_index], rows[right_index]):
                    union_find.union(left_index, right_index)

    buckets: dict[tuple[tuple[str, ...], str], list[int]] = defaultdict(list)
    for index, row in enumerate(rows):
        description = canonical(row.descripcion)
        prefix = " ".join(description.split()[:2])
        buckets[(measurement_signature(row.descripcion), prefix)].append(index)

    for indices in buckets.values():
        if len(indices) > 80:
            continue
        for position, left_index in enumerate(indices):
            for right_index in indices[position + 1 :]:
                if union_find.find(left_index) != union_find.find(right_index) and should_merge(rows[left_index], rows[right_index]):
                    union_find.union(left_index, right_index)

    groups: dict[int, list[SourceRow]] = defaultdict(list)
    for index, row in enumerate(rows):
        groups[union_find.find(index)].append(row)
    return list(groups.values())


def best_record(rows: Iterable[SourceRow], require_part: bool = False) -> SourceRow | None:
    candidates = [row for row in rows if not require_part or row.numero_parte]
    if not candidates:
        return None
    return max(
        candidates,
        key=lambda row: (
            bool(row.stock is not None),
            row.snapshot_date or dt.date.min,
            row.source_priority,
            row.completeness,
            len(row.descripcion),
        ),
    )


def choose_mode(rows: Iterable[SourceRow], field_name: str, fallback: str = "") -> str:
    values = [normalize(getattr(row, field_name)) for row in rows if normalize(getattr(row, field_name))]
    if not values:
        return fallback
    counts = Counter(canonical(value) for value in values)
    winner = max(counts, key=lambda key: (counts[key], len(key)))
    matching = [value for value in values if canonical(value) == winner]
    return max(matching, key=len)


def aggregate_products(groups: list[list[SourceRow]], image_dir: Path) -> tuple[list[Product], dict[str, bytes]]:
    products: list[Product] = []
    images_to_write: dict[str, bytes] = {}

    for group in groups:
        per_warehouse: dict[str, list[SourceRow]] = defaultdict(list)
        for row in group:
            per_warehouse[row.warehouse].append(row)

        selected_by_warehouse: list[SourceRow] = []
        conflicts: list[str] = []
        for warehouse, candidates in per_warehouse.items():
            selected = best_record(candidates)
            assert selected is not None
            selected_by_warehouse.append(selected)
            known_values = sorted({candidate.stock for candidate in candidates if candidate.stock is not None})
            if len(known_values) > 1:
                conflicts.append(f"{warehouse}: stocks observados {', '.join(map(str, known_values))}; se tomó {selected.stock}")

        metadata = best_record(group)
        assert metadata is not None
        part_record = best_record(group, require_part=True)
        numero_parte = part_record.numero_parte if part_record else ""
        descripcion = choose_mode(group, "descripcion", metadata.descripcion)
        marca = choose_mode(group, "marca", metadata.marca)
        categoria = clean_category(choose_mode(group, "categoria", metadata.categoria), metadata.warehouse)
        warehouses = sorted(per_warehouse)
        known_rows = [row for row in selected_by_warehouse if row.stock is not None]
        stock = sum(int(row.stock or 0) for row in known_rows)

        image_rows = [row for row in group if row.image is not None]
        image_record = max(
            image_rows,
            key=lambda row: (
                row.snapshot_date or dt.date.min,
                row.source_priority,
                row.image.display_area if row.image else 0,
            ),
            default=None,
        )
        fotografia: str | None = None
        if image_record and image_record.image:
            extension_name = f"{image_record.image.content_hash[:20]}.jpg"
            fotografia = f"materiales/base-2026/{extension_name}"
            images_to_write.setdefault(extension_name, image_record.image.data)

        identity = "|".join((canonical(numero_parte), canonical(descripcion), canonical(marca)))
        source_key = hashlib.sha1(identity.encode("utf-8")).hexdigest()
        notes = list(conflicts)
        if not known_rows:
            notes.append("Stock no informado en el Excel; se importará en 0")
        if len(group) > 1:
            notes.append(f"Se consolidaron {len(group)} renglones fuente")

        products.append(
            Product(
                source_key=source_key,
                numero_parte=numero_parte,
                descripcion=descripcion,
                marca=marca,
                categoria=categoria,
                almacen="; ".join(warehouses),
                stock=stock,
                stock_known=bool(known_rows),
                fotografia=fotografia,
                sources=sorted(group, key=lambda row: (row.sheet, row.row)),
                notes=notes,
            )
        )

    products.sort(key=lambda product: (canonical(product.categoria), canonical(product.descripcion), canonical(product.numero_parte)))
    return products, images_to_write


def save_optimized_jpeg(data: bytes, target: Path) -> None:
    with Image.open(BytesIO(data)) as image:
        image = ImageOps.exif_transpose(image)
        if image.mode in {"RGBA", "LA"} or (image.mode == "P" and "transparency" in image.info):
            rgba = image.convert("RGBA")
            background = Image.new("RGBA", rgba.size, "white")
            background.alpha_composite(rgba)
            image = background.convert("RGB")
        else:
            image = image.convert("RGB")
        image.thumbnail((960, 960), Image.Resampling.LANCZOS)
        target.parent.mkdir(parents=True, exist_ok=True)
        image.save(target, "JPEG", quality=74, optimize=True, progressive=True)


def sql_string(value: str | None) -> str:
    if value is None or value == "":
        return "NULL"
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def generate_sql(products: list[Product], source: Path, target: Path) -> None:
    file_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    lines = [
        "-- Importación controlada de BASE DE DATOS 2026.xlsx",
        "-- Ejecutar el archivo completo en DBeaver sobre la base correcta.",
        "-- No elimina materiales existentes. Consolida duplicados del Excel y registra ajustes de stock.",
        "-- @actualizar_stock_existente = 1 reemplaza el stock de coincidencias exactas con el del Excel.",
        "SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci;",
        "SET @actualizar_stock_existente := 1;",
        f"SET @clave_importacion := '{IMPORT_KEY}';",
        "START TRANSACTION;",
        "",
        "CREATE TABLE IF NOT EXISTS `inventario_importaciones_manuales` (",
        "  `clave` varchar(100) NOT NULL,",
        "  `archivo` varchar(255) NOT NULL,",
        "  `hash_archivo` char(64) NOT NULL,",
        "  `registros` int unsigned NOT NULL,",
        "  `ejecutado_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP,",
        "  `detalles` json NULL,",
        "  PRIMARY KEY (`clave`)",
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;",
        "",
        "SET @ya_importado := (SELECT COUNT(*) FROM `inventario_importaciones_manuales` WHERE `clave` = @clave_importacion);",
        "DROP TEMPORARY TABLE IF EXISTS `tmp_base_datos_2026`;",
        "CREATE TEMPORARY TABLE `tmp_base_datos_2026` (",
        "  `source_key` char(40) NOT NULL PRIMARY KEY,",
        "  `categoria` varchar(255) NOT NULL,",
        "  `almacen` varchar(255) NOT NULL,",
        "  `numero_parte` varchar(255) NULL,",
        "  `unidad` varchar(80) NOT NULL,",
        "  `descripcion` text NOT NULL,",
        "  `marca` varchar(255) NULL,",
        "  `stock` int unsigned NOT NULL,",
        "  `stock_conocido` tinyint(1) NOT NULL,",
        "  `fotografia` varchar(255) NULL,",
        "  `material_id` bigint unsigned NULL,",
        "  `existia` tinyint(1) NOT NULL DEFAULT 0",
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;",
        "",
    ]

    for start in range(0, len(products), 100):
        batch = products[start : start + 100]
        lines.append(
            "INSERT INTO `tmp_base_datos_2026` (`source_key`,`categoria`,`almacen`,`numero_parte`,`unidad`,`descripcion`,`marca`,`stock`,`stock_conocido`,`fotografia`) VALUES"
        )
        values = []
        for product in batch:
            values.append(
                "(" + ",".join(
                    (
                        sql_string(product.source_key),
                        sql_string(product.categoria),
                        sql_string(product.almacen),
                        sql_string(product.numero_parte),
                        "'pza'",
                        sql_string(product.descripcion),
                        sql_string(product.marca),
                        str(product.stock),
                        "1" if product.stock_known else "0",
                        sql_string(product.fotografia),
                    )
                ) + ")"
            )
        lines.append(",\n".join(values) + ";")
        lines.append("")

    lines.extend(
        [
            "-- Sincroniza las categorías del Excel con el catálogo editable de la aplicación.",
            "-- Se ejecuta incluso si la carga de existencias ya se había realizado.",
            "INSERT IGNORE INTO `material_categories` (`nombre`,`descripcion`,`activa`,`created_at`,`updated_at`)",
            "SELECT DISTINCT TRIM(s.`categoria`), NULL, 1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP",
            "FROM `tmp_base_datos_2026` s",
            "WHERE TRIM(s.`categoria`) <> '' AND UPPER(TRIM(s.`categoria`)) NOT LIKE 'EQUIPO%';",
            "",
            "-- Primero se busca una coincidencia exacta por número de parte y descripción.",
            "DROP TEMPORARY TABLE IF EXISTS `tmp_base_datos_2026_coincidencias`;",
            "CREATE TEMPORARY TABLE `tmp_base_datos_2026_coincidencias` (",
            "  `source_key` char(40) NOT NULL PRIMARY KEY,",
            "  `material_id` bigint unsigned NOT NULL",
            ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;",
            "INSERT INTO `tmp_base_datos_2026_coincidencias` (`source_key`,`material_id`)",
            "SELECT s2.`source_key`, MIN(m.`id`) AS `material_id`",
            "FROM `tmp_base_datos_2026` s2",
            "JOIN `materials` m ON COALESCE(m.`es_plantilla_equipo`, 0) = 0",
            "  AND s2.`numero_parte` IS NOT NULL",
            "  AND UPPER(TRIM(m.`numero_parte`)) = UPPER(TRIM(s2.`numero_parte`))",
            "  AND UPPER(TRIM(m.`descripcion`)) = UPPER(TRIM(s2.`descripcion`))",
            "WHERE @ya_importado = 0",
            "GROUP BY s2.`source_key`;",
            "UPDATE `tmp_base_datos_2026` s",
            "JOIN `tmp_base_datos_2026_coincidencias` x ON x.`source_key` = s.`source_key`",
            "SET s.`material_id` = x.`material_id`, s.`existia` = 1",
            "WHERE @ya_importado = 0;",
            "",
            "-- Para los restantes se exige descripción exacta y marca compatible.",
            "TRUNCATE TABLE `tmp_base_datos_2026_coincidencias`;",
            "INSERT INTO `tmp_base_datos_2026_coincidencias` (`source_key`,`material_id`)",
            "SELECT s2.`source_key`, MIN(m.`id`) AS `material_id`",
            "FROM `tmp_base_datos_2026` s2",
            "JOIN `materials` m ON COALESCE(m.`es_plantilla_equipo`, 0) = 0",
            "  AND UPPER(TRIM(m.`descripcion`)) = UPPER(TRIM(s2.`descripcion`))",
            "  AND (s2.`marca` IS NULL OR m.`marca` IS NULL OR UPPER(TRIM(m.`marca`)) = UPPER(TRIM(s2.`marca`)))",
            "WHERE @ya_importado = 0 AND s2.`material_id` IS NULL",
            "GROUP BY s2.`source_key`;",
            "UPDATE `tmp_base_datos_2026` s",
            "JOIN `tmp_base_datos_2026_coincidencias` x ON x.`source_key` = s.`source_key`",
            "SET s.`material_id` = x.`material_id`, s.`existia` = 1",
            "WHERE @ya_importado = 0 AND s.`material_id` IS NULL;",
            "",
            "DROP TEMPORARY TABLE IF EXISTS `tmp_base_datos_2026_ajustes`;",
            "CREATE TEMPORARY TABLE `tmp_base_datos_2026_ajustes` AS",
            "SELECT s.`material_id`, m.`stock` AS `stock_anterior`, s.`stock` AS `stock_nuevo`",
            "FROM `tmp_base_datos_2026` s",
            "JOIN `materials` m ON m.`id` = s.`material_id`",
            "WHERE @ya_importado = 0 AND @actualizar_stock_existente = 1 AND s.`stock_conocido` = 1 AND m.`stock` <> s.`stock`;",
            "",
            "UPDATE `materials` m",
            "JOIN `tmp_base_datos_2026` s ON s.`material_id` = m.`id`",
            "SET m.`categoria` = COALESCE(NULLIF(m.`categoria`, ''), s.`categoria`),",
            "    m.`almacen` = s.`almacen`,",
            "    m.`numero_parte` = COALESCE(NULLIF(m.`numero_parte`, ''), s.`numero_parte`),",
            "    m.`unidad` = COALESCE(NULLIF(m.`unidad`, ''), s.`unidad`),",
            "    m.`marca` = COALESCE(NULLIF(m.`marca`, ''), s.`marca`),",
            "    m.`fotografia` = COALESCE(NULLIF(m.`fotografia`, ''), s.`fotografia`),",
            "    m.`stock` = CASE WHEN @actualizar_stock_existente = 1 AND s.`stock_conocido` = 1 THEN s.`stock` ELSE m.`stock` END,",
            "    m.`updated_at` = CURRENT_TIMESTAMP",
            "WHERE @ya_importado = 0;",
            "",
            "INSERT INTO `materials` (`categoria`,`almacen`,`numero_parte`,`codigo_barras`,`unidad`,`descripcion`,`apodo`,`es_plantilla_equipo`,`marca`,`proveedor`,`stock`,`stock_minimo`,`stock_maximo`,`costo_unitario`,`moneda`,`fotografia`,`created_at`,`updated_at`)",
            "SELECT s.`categoria`, s.`almacen`, s.`numero_parte`, NULL, s.`unidad`, s.`descripcion`, NULL, 0, s.`marca`, NULL,",
            "       CASE WHEN s.`stock_conocido` = 1 THEN s.`stock` ELSE 0 END, 0, 0, 0, 'MXN', s.`fotografia`, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP",
            "FROM `tmp_base_datos_2026` s",
            "WHERE @ya_importado = 0 AND s.`material_id` IS NULL;",
            "",
            "-- Recupera el id de cada alta para dejar trazabilidad de la carga inicial.",
            "TRUNCATE TABLE `tmp_base_datos_2026_coincidencias`;",
            "INSERT INTO `tmp_base_datos_2026_coincidencias` (`source_key`,`material_id`)",
            "SELECT s2.`source_key`, MIN(m.`id`) AS `material_id`",
            "FROM `tmp_base_datos_2026` s2",
            "JOIN `materials` m ON UPPER(TRIM(m.`descripcion`)) = UPPER(TRIM(s2.`descripcion`))",
            "  AND (s2.`numero_parte` IS NULL OR UPPER(TRIM(m.`numero_parte`)) = UPPER(TRIM(s2.`numero_parte`)))",
            "  AND (s2.`marca` IS NULL OR m.`marca` IS NULL OR UPPER(TRIM(m.`marca`)) = UPPER(TRIM(s2.`marca`)))",
            "WHERE @ya_importado = 0 AND s2.`material_id` IS NULL",
            "GROUP BY s2.`source_key`;",
            "UPDATE `tmp_base_datos_2026` s",
            "JOIN `tmp_base_datos_2026_coincidencias` x ON x.`source_key` = s.`source_key`",
            "SET s.`material_id` = x.`material_id`",
            "WHERE @ya_importado = 0 AND s.`material_id` IS NULL;",
            "",
            "INSERT INTO `material_movimientos` (`material_id`,`user_id`,`tipo`,`cantidad`,`stock_anterior`,`stock_nuevo`,`codigo_barras`,`referencia`,`motivo`,`evidencia_foto`,`proveedor`,`costo_unitario`,`created_at`,`updated_at`)",
            "SELECT a.`material_id`, NULL, 'ajuste', ABS(a.`stock_nuevo` - a.`stock_anterior`), a.`stock_anterior`, a.`stock_nuevo`, NULL,",
            "       'BASE DE DATOS 2026', 'Ajuste inicial desde la base de datos física 2026', NULL, NULL, NULL, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP",
            "FROM `tmp_base_datos_2026_ajustes` a",
            "WHERE @ya_importado = 0;",
            "",
            "INSERT INTO `material_movimientos` (`material_id`,`user_id`,`tipo`,`cantidad`,`stock_anterior`,`stock_nuevo`,`codigo_barras`,`referencia`,`motivo`,`evidencia_foto`,`proveedor`,`costo_unitario`,`created_at`,`updated_at`)",
            "SELECT s.`material_id`, NULL, 'entrada', s.`stock`, 0, s.`stock`, NULL,",
            "       'BASE DE DATOS 2026', 'Carga inicial desde la base de datos física 2026', NULL, NULL, NULL, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP",
            "FROM `tmp_base_datos_2026` s",
            "WHERE @ya_importado = 0 AND s.`existia` = 0 AND s.`stock_conocido` = 1 AND s.`stock` > 0 AND s.`material_id` IS NOT NULL;",
            "",
            "INSERT INTO `inventario_importaciones_manuales` (`clave`,`archivo`,`hash_archivo`,`registros`,`detalles`)",
            f"SELECT @clave_importacion, {sql_string(source.name)}, '{file_hash}', {len(products)}, JSON_OBJECT('stock_existente_actualizado', @actualizar_stock_existente)",
            "WHERE @ya_importado = 0;",
            "",
            "COMMIT;",
            "",
            "-- Resultado de control. Si ya se había ejecutado, no vuelve a modificar existencias.",
            "SELECT @ya_importado AS `ya_estaba_importado`, COUNT(*) AS `renglones_preparados`,",
            "       SUM(`stock_conocido`) AS `con_stock_conocido`, SUM(`stock`) AS `piezas_fuente`",
            "FROM `tmp_base_datos_2026`;",
            "SELECT `clave`,`archivo`,`hash_archivo`,`registros`,`ejecutado_at` FROM `inventario_importaciones_manuales` WHERE `clave` = @clave_importacion;",
            "SELECT COUNT(*) AS `categorias_en_catalogo` FROM `material_categories` WHERE UPPER(TRIM(`nombre`)) NOT LIKE 'EQUIPO%';",
            "",
        ]
    )
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text("\n".join(lines), encoding="utf-8")


def write_reports(
    products: list[Product],
    raw_rows: list[SourceRow],
    skipped: list[dict[str, Any]],
    source: Path,
    report_csv: Path,
    possible_csv: Path,
    summary_json: Path,
    image_count: int,
) -> None:
    report_csv.parent.mkdir(parents=True, exist_ok=True)
    with report_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "numero_parte",
                "descripcion",
                "marca",
                "categoria",
                "almacen",
                "stock",
                "stock_conocido",
                "fotografia",
                "renglones_consolidados",
                "fuentes_excel",
                "observaciones",
            ],
        )
        writer.writeheader()
        for product in products:
            writer.writerow(
                {
                    "numero_parte": product.numero_parte,
                    "descripcion": product.descripcion,
                    "marca": product.marca,
                    "categoria": product.categoria,
                    "almacen": product.almacen,
                    "stock": product.stock,
                    "stock_conocido": "Sí" if product.stock_known else "No",
                    "fotografia": product.fotografia or "",
                    "renglones_consolidados": len(product.sources),
                    "fuentes_excel": "; ".join(row.source_ref for row in product.sources),
                    "observaciones": "; ".join(product.notes),
                }
            )

    possible_pairs: list[dict[str, Any]] = []
    comparison_buckets: dict[str, list[Product]] = defaultdict(list)
    for product in products:
        description = canonical(product.descripcion)
        comparison_buckets[(description.split() or [""])[0]].append(product)

    for candidates in comparison_buckets.values():
        for left_index, left in enumerate(candidates):
            left_desc = canonical(left.descripcion)
            left_signature = measurement_signature(left.descripcion)
            for right in candidates[left_index + 1 :]:
                right_desc = canonical(right.descripcion)
                if abs(len(left_desc) - len(right_desc)) > max(12, int(max(len(left_desc), len(right_desc)) * 0.18)):
                    continue
                if not compatible_brand(left.marca, right.marca):
                    continue
                ratio = SequenceMatcher(None, left_desc, right_desc).ratio()
                if ratio < 0.90:
                    continue
                right_signature = measurement_signature(right.descripcion)
                reason = "Medidas/modelos diferentes" if left_signature != right_signature else "Similitud insuficiente para fusionar automáticamente"
                possible_pairs.append(
                    {
                        "similitud": round(ratio * 100, 1),
                        "pieza_1": left.descripcion,
                        "no_parte_1": left.numero_parte,
                        "pieza_2": right.descripcion,
                        "no_parte_2": right.numero_parte,
                        "razon_no_fusion": reason,
                    }
                )
    possible_pairs.sort(key=lambda item: item["similitud"], reverse=True)
    with possible_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        fields = ["similitud", "pieza_1", "no_parte_1", "pieza_2", "no_parte_2", "razon_no_fusion"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(possible_pairs)

    summary = {
        "source": str(source),
        "source_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "raw_product_rows": len(raw_rows),
        "consolidated_products": len(products),
        "duplicate_rows_consolidated": len(raw_rows) - len(products),
        "products_with_known_stock": sum(product.stock_known for product in products),
        "products_without_known_stock": sum(not product.stock_known for product in products),
        "known_stock_total": sum(product.stock for product in products if product.stock_known),
        "products_with_photo": sum(bool(product.fotografia) for product in products),
        "unique_optimized_images": image_count,
        "possible_duplicates_not_merged": len(possible_pairs),
        "skipped_sheets_or_rows": skipped,
        "warehouses": dict(sorted(Counter(row.warehouse for row in raw_rows).items())),
    }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def discover_default_source() -> Path:
    matches = list(Path("D:/LUGARTH").rglob(DEFAULT_SOURCE_NAME))
    if not matches:
        raise FileNotFoundError(f"No se encontró {DEFAULT_SOURCE_NAME} dentro de D:/LUGARTH")
    return matches[0]


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Genera la importación controlada del inventario físico 2026.")
    parser.add_argument("--source", type=Path, default=None)
    parser.add_argument("--sql", type=Path, default=Path("database/sql/importar_base_datos_2026.sql"))
    parser.add_argument("--report", type=Path, default=Path("database/sql/base_datos_2026_revision.csv"))
    parser.add_argument("--possible-duplicates", type=Path, default=Path("database/sql/base_datos_2026_posibles_duplicados.csv"))
    parser.add_argument("--summary", type=Path, default=Path("database/sql/base_datos_2026_resumen.json"))
    parser.add_argument("--images", type=Path, default=Path("storage/app/public/materiales/base-2026"))
    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    source = args.source or discover_default_source()
    workbook = openpyxl.load_workbook(source, data_only=True)
    raw_rows, skipped = read_source_rows(workbook)
    groups = cluster_rows(raw_rows)
    products, images_to_write = aggregate_products(groups, args.images)

    args.images.mkdir(parents=True, exist_ok=True)
    expected_files = set(images_to_write)
    for existing in args.images.glob("*.jpg"):
        if existing.name not in expected_files:
            existing.unlink()
    for filename, data in images_to_write.items():
        target = args.images / filename
        if not target.is_file():
            save_optimized_jpeg(data, target)

    generate_sql(products, source, args.sql)
    write_reports(
        products,
        raw_rows,
        skipped,
        source,
        args.report,
        args.possible_duplicates,
        args.summary,
        len(images_to_write),
    )

    print(f"Renglones leídos: {len(raw_rows)}")
    print(f"Piezas consolidadas: {len(products)}")
    print(f"Fotos optimizadas únicas: {len(images_to_write)}")
    print(f"SQL: {args.sql.resolve()}")
    print(f"Reporte: {args.report.resolve()}")


if __name__ == "__main__":
    main()
